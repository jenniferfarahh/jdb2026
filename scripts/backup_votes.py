#!/usr/bin/env python3
"""
JDB 2026 — Backup des votes
Génère un fichier Excel (.xlsx) avec 2 onglets dans ~/Downloads :
  - Votes      : un votant par ligne avec tous ses choix
  - Résultats  : classement projets + OBNLs triés par points

Usage:
    bash /Users/jen/Desktop/jdb2026/scripts/backup.sh
"""

import psycopg2, re, sys
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_URL       = "postgresql://neondb_owner:npg_XhT4uRDQiaI9@ep-late-tree-amb2gnev-pooler.c-5.us-east-1.aws.neon.tech/neondb?sslmode=require"
POOL_PROJETS = 35000
POOL_ONBS    = 5000
DOWNLOADS    = Path.home() / "Downloads"
REPO         = Path(__file__).parent.parent

# ── Styles ────────────────────────────────────────────────────────────────────
BLUE   = "FF2563EB"
TEAL   = "FF2ABFC4"
WHITE  = "FFFFFFFF"
GRAY   = "FFF1F5F9"
DARK   = "FF1E293B"
GOLD   = "FFFBBF24"

def h(wb, text, color=BLUE):
    """Bold white header cell with colored background."""
    f = Font(bold=True, color=WHITE, size=10)
    fill = PatternFill("solid", fgColor=color)
    al = Alignment(horizontal="center", vertical="center")
    return {"font": f, "fill": fill, "alignment": al}

def style_header_row(ws, row, col_start, col_end, color=BLUE):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_section_title(ws, row, col, text, color=BLUE):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22

def autofit(ws, min_w=8, max_w=40):
    for col in ws.columns:
        width = min_w
        for cell in col:
            if cell.value:
                width = min(max_w, max(width, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = width

# ── Load static data ──────────────────────────────────────────────────────────
def load_projects():
    content = (REPO / "data" / "projets.ts").read_text()
    out = {}
    for m in re.finditer(r"\{\s*id:'([^']+)',\s*name:'([^']+)'.*?montant:(\d+)", content, re.DOTALL):
        out[m.group(1)] = {"name": m.group(2), "montant": int(m.group(3))}
    return out

def load_ongs():
    content = (REPO / "data" / "ong.ts").read_text()
    out = {}
    for m in re.finditer(r"id:\s*['\"]([^'\"]+)['\"].*?name:\s*['\"]([^'\"]+)['\"]", content, re.DOTALL):
        out[m.group(1)] = m.group(2)
    return out

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("🔌 Connexion à Neon Postgres…")
    try:
        conn = psycopg2.connect(DB_URL)
    except Exception as e:
        print(f"❌ Erreur de connexion : {e}")
        sys.exit(1)

    cur = conn.cursor()
    projects = load_projects()
    ongs     = load_ongs()
    ts       = datetime.now().strftime("%Y-%m-%d_%H%M")

    cur.execute("SELECT id, prenom, nom, email, promo_type, voter_category, voted_at FROM vote_sessions ORDER BY voted_at")
    sessions = cur.fetchall()

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 1 — Votes
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Votes"
    ws1.freeze_panes = "A2"

    headers1 = [
        "Prénom","Nom","Email","Promo","Catégorie","Heure du vote",
        "Projet #1","Poids","Projet #2","Poids","Projet #3","Poids",
        "Projet #4","Poids","Projet #5","Poids",
        "OBNL #1","Poids","OBNL #2","Poids","OBNL #3","Poids",
    ]
    ws1.append(headers1)
    style_header_row(ws1, 1, 1, len(headers1), BLUE)
    ws1.row_dimensions[1].height = 20

    alt = False
    for sid, prenom, nom, email, promo, cat, voted_at in sessions:
        cur.execute("SELECT rank, project_id, weight FROM project_votes WHERE session_id=%s ORDER BY rank", (sid,))
        pvotes = {r: (p, w) for r, p, w in cur.fetchall()}
        cur.execute("SELECT rank, ong_id, weight FROM ong_votes WHERE session_id=%s ORDER BY rank", (sid,))
        ovotes = {r: (o, w) for r, o, w in cur.fetchall()}

        row = [prenom, nom, email or "", promo, cat,
               voted_at.strftime("%d/%m/%Y %H:%M") if voted_at else ""]
        for i in range(1, 6):
            if i in pvotes:
                pid, w = pvotes[i]
                row += [projects.get(pid, {}).get("name", pid), w]
            else:
                row += ["", ""]
        for i in range(1, 4):
            if i in ovotes:
                oid, w = ovotes[i]
                row += [ongs.get(oid, oid), w]
            else:
                row += ["", ""]

        ws1.append(row)
        if alt:
            for col in range(1, len(headers1) + 1):
                ws1.cell(ws1.max_row, col).fill = PatternFill("solid", fgColor=GRAY)
        alt = not alt

    autofit(ws1)

    # ══════════════════════════════════════════════════════════════════════════
    # ONGLET 2 — Résultats
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Résultats")

    # Fetch scores
    cur.execute("SELECT project_id, SUM(weight)::int, COUNT(*)::int FROM project_votes GROUP BY project_id ORDER BY SUM(weight) DESC")
    proj_scores = cur.fetchall()
    total_proj_pts = sum(r[1] for r in proj_scores)

    cur.execute("SELECT ong_id, SUM(weight)::int, COUNT(*)::int FROM ong_votes GROUP BY ong_id ORDER BY SUM(weight) DESC")
    ong_scores = cur.fetchall()
    total_ong_pts = sum(r[1] for r in ong_scores)

    cur.execute("SELECT voter_category, COUNT(*) FROM vote_sessions GROUP BY voter_category ORDER BY voter_category")
    by_cat = cur.fetchall()
    total_voters = len(sessions)

    r = 1

    # Stats
    style_section_title(ws2, r, 1, "STATISTIQUES", DARK)
    ws2.merge_cells(f"A{r}:B{r}")
    r += 1
    ws2.cell(r, 1, "Total votants").font = Font(bold=True)
    ws2.cell(r, 2, total_voters)
    r += 1
    for cat, count in by_cat:
        ws2.cell(r, 1, f"  {cat}")
        ws2.cell(r, 2, count)
        r += 1

    r += 1

    # Projets header
    style_section_title(ws2, r, 1, f"CLASSEMENT PROJETS  —  Pool : {POOL_PROJETS:,} €".replace(",", " "), BLUE)
    ws2.merge_cells(f"A{r}:G{r}")
    r += 1

    proj_cols = ["#","Projet","Points","Part (%)","Alloué (€)","Demandé (€)","Nb votants"]
    ws2.append(proj_cols)
    style_header_row(ws2, r, 1, len(proj_cols), BLUE)
    ws2.row_dimensions[r].height = 20
    r += 1

    medals = {1: "🥇", 2: "🥈", 3: "🥉"}
    for i, (pid, pts, nb) in enumerate(proj_scores, 1):
        meta  = projects.get(pid, {})
        part  = pts / total_proj_pts if total_proj_pts else 0
        alloc = round(part * POOL_PROJETS)
        label = f"{medals.get(i,'')} {i}".strip()
        ws2.append([label, meta.get("name", pid), pts, f"{part*100:.1f}%", alloc, meta.get("montant", 0), nb])
        if i <= 3:
            ws2.cell(r, 1).font = Font(bold=True)
            ws2.cell(r, 2).font = Font(bold=True)
            ws2.cell(r, 5).font = Font(bold=True, color=BLUE)
        r += 1

    r += 1

    # OBNLs header
    style_section_title(ws2, r, 1, f"CLASSEMENT OBNLs  —  Pool : {POOL_ONBS:,} €".replace(",", " "), TEAL)
    ws2.merge_cells(f"A{r}:F{r}")
    r += 1

    ong_cols = ["#","OBNL","Points","Part (%)","Alloué (€)","Nb votants"]
    ws2.append(ong_cols)
    style_header_row(ws2, r, 1, len(ong_cols), TEAL)
    ws2.row_dimensions[r].height = 20
    r += 1

    for i, (oid, pts, nb) in enumerate(ong_scores, 1):
        name  = ongs.get(oid, oid)
        part  = pts / total_ong_pts if total_ong_pts else 0
        alloc = round(part * POOL_ONBS)
        label = f"{medals.get(i,'')} {i}".strip()
        ws2.append([label, name, pts, f"{part*100:.1f}%", alloc, nb])
        if i <= 3:
            ws2.cell(r, 1).font = Font(bold=True)
            ws2.cell(r, 2).font = Font(bold=True)
            ws2.cell(r, 5).font = Font(bold=True, color=TEAL)
        r += 1

    autofit(ws2)

    # ── Save ──────────────────────────────────────────────────────────────────
    path = DOWNLOADS / f"JDB2026_Export_{ts}.xlsx"
    wb.save(path)
    print(f"✅ {total_voters} votes  →  {path}")
    print(f"   Onglet 'Votes'     : {len(sessions)} lignes")
    print(f"   Onglet 'Résultats' : {len(proj_scores)} projets · {len(ong_scores)} OBNLs")
    conn.close()

if __name__ == "__main__":
    main()
